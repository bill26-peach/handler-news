import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def json_to_excel(json_str, output_file):
    parsed_data = [json.loads(item) for item in json_str]
    header_mapping = {
        "catm": "报告时间", "site_name": "媒体", "cntt": "内容",
        "titl": "新闻标题", "sentiment": "正负中立", "color": "蓝绿", "keyword": "关键字",
        "type": "类型(热点追踪、台媒焦点、重点关注、经贸资讯、港澳动态、其他)"
    }
    all_columns = ["来源", "期数", "报告时间", "类型(热点追踪、台媒焦点、重点关注、经贸资讯、港澳动态、其他)",
                   "新闻标题", "地区", "内容", "媒体", "蓝绿", "正负中立", "关键字", "舆情等级", "日周月报",
                   "台胞台企", "级别"]
    df = pd.DataFrame(parsed_data)
    df.rename(columns=header_mapping, inplace=True)
    df = df.reindex(columns=all_columns)
    df["来源"] = df["媒体"]
    df.fillna("", inplace=True)
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions["G"].width = 50

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output_file)
    print(f"Excel文件已保存到: {output_file}")



